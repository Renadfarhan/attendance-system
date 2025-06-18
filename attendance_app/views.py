import os
from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from datetime import datetime

def mark_attendance(request):
    if request.method == "POST":
        name = request.POST.get('name')
        timestamp = request.POST.get('timestame')
        if name:
            # Excel file path
            file_path = 'attendance_records.xlsx'

            # Create Excel file if it doesn't exist
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.append(['Name', 'Timestamp'])  # Header
                wb.save(file_path)

            # Load and append new data
            wb = load_workbook(file_path)
            ws = wb.active
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.append([name, timestamp])
            wb.save(file_path)

            return render(request, 'attendance_app/success.html', {'name': name, 'timestamp': timestamp})

    return render(request, 'attendance_app/attendance_form.html')
